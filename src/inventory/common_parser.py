import re
import sys
import os
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from openpyxl import load_workbook

from inventory.varnames import ColNames as C
from api_integrations.sharepoint_client import SharePointClient
invoc = SharePointClient()

def get_all_csv_files_in_directory(directory_path):
    if not os.path.isdir(directory_path):
        raise FileNotFoundError(f"{directory_path} is not a valid directory.")

    return [
        os.path.join(directory_path, f)
        for f in os.listdir(directory_path)
        if os.path.isfile(os.path.join(directory_path, f)) and f.lower().endswith('.csv')
    ]

def read_files(temp_paths, update_from_sharepoint):
    config = invoc.read_json("config/config.json")
    inventory_df = invoc.read_excel('INVENTARIO/INVENTARIO.xlsx')
    if update_from_sharepoint:
        po_df = invoc.read_excel(f'RECIBOS/{update_from_sharepoint}.xlsx')
        po_type = 'update'
    else:
        po_dfs = [pd.read_csv(po_path) for po_path in temp_paths]
        # po_read_path = '../../files/inventory/drag_and_drop'  ##for local debugging
        # po_files = get_all_csv_files_in_directory(po_read_path)
        # po_dfs = [pd.read_csv(po_path) for po_path in po_files]
        po_df = pd.concat(po_dfs)
        po_type = auto_assign_po_type(po_df)
    po_df['Nombre Tienda'] = assign_store_name(po_df, po_type)
    cols_rename = config[f'{po_type.lower()}_rename']
    po_df = po_df.rename(columns=cols_rename)
    if po_type != 'receipt':
        matching_cols = [C.WAREHOUSE_CODE, C.SKU, C.UPC, C.STYLE]
        matching_column = auto_assign_matching_column(po_df, matching_cols)
        cols = [matching_column] + [x for x in cols_rename.values() if x not in matching_cols]
    else:
        matching_column = 'index'
        cols = list(cols_rename.values())
    return ((po_df.reset_index()
             .sort_values(by=matching_column)[cols]
             .reset_index(drop=True)),
            inventory_df, config, po_type, matching_column)


def assign_store_name(po_df, po_type):
    store_mapping = invoc.read_csv(f"config/tiendas_{po_type.lower()}.csv")
    po_df = po_df.merge(store_mapping, how='left')
    return po_df['Nombre Tienda'].fillna("NotFound")


def auto_assign_po_type(df):
    if '# Prov' in df.columns:
        return 'liverpool'
    elif 'Num. Prov' in df.columns:
        return 'suburbia'
    elif 'RD' in df.columns:
        return 'receipt'
    return 'interno'

def auto_assign_matching_column(df, lst):
    for col in lst:
        if col in df.columns:
            return col
    sys.exit("Error: File must contain a matching column.")

def assign_warehouse_codes_from_column_and_update_inventory(po, inventory, column, log_id):
    split_inventory = split_df_by_column(inventory.copy(), column)
    po_missing = po.loc[(po[C.DELIVERED] == 0)].merge(
            split_inventory[1],
            on=[column], how='left')
    po_original_cols = po.columns
    it = 0
    po_wh = [po_missing]
    updated_inv = [split_inventory[0]]
    for inventory_wh in split_inventory[1:]:
        it += 1
        po, to_deliver = assign_warehouse_codes(po, inventory_wh, column)
        po_wh.append(po.loc[po[C.DELIVERED] > 0].copy())
        update_inventory(inventory_wh, po, updated_inv, log_id)
        po[C.DELIVERED] = to_deliver
        po = po.loc[po[C.DELIVERED] > 0, po_original_cols]
        if len(po) == 0:
            break
    po = pd.concat(po_wh)
    po = split_ordered_quantity_by_warehouse_codes(po, column)
    updated_inv_lst = updated_inv + split_inventory[it+1:]
    updated_inv = concat_inv_lst(updated_inv_lst)
    for df in [po, updated_inv]:
        convert_cols_to_int(df, [C.WAREHOUSE_CODE, C.UPC, C.SKU])
    return po.sort_values([C.STORE_ID, column]).reset_index(drop=True), updated_inv

def convert_cols_to_int(df, cols):
    for col in cols:
        if col in df.columns:
            df[col] = df[col].astype(int)

def concat_inv_lst(dfs):
    df = pd.concat(dfs)
    # df = df.sort_values(
    #     by=['RD', 'WAREHOUSE_CODE'],
    #     key=lambda col: col.map(sort_rd) if col.name == 'code' else col,
    #     ignore_index=True
    # )
    df.sort_index(inplace=True)
    return df


def assign_warehouse_codes(po, inventory_wh, column):
    po = po.merge(
        inventory_wh,
        on=[column], how='left')
    delivered_cs = po.groupby(column)[C.DELIVERED].cumsum()
    missing = (po[C.INVENTORY] - delivered_cs).clip(upper=0).fillna(0)
    delivered_i = (po[C.DELIVERED] + missing).clip(lower=0)
    to_deliver = po[C.DELIVERED] - delivered_i
    po[C.DELIVERED] = delivered_i
    return po, to_deliver


def split_df_by_column(df, column):
    column0 = df[df[column] == 0]
    df = df.drop(index=column0.index)
    split_dfs = [column0]
    while not df.empty:
        unique_df = df.drop_duplicates(subset=[column], keep='first')
        split_dfs.append(unique_df)
        df = df.drop(index=unique_df.index)
    return split_dfs

def update_inventory(inventory_wh, po, updated_inv, log_id):
    inventory_wh = inventory_wh.join(po.groupby([C.WAREHOUSE_CODE])[C.DELIVERED].sum(), on=C.WAREHOUSE_CODE)
    inventory_wh[C.DELIVERED] = inventory_wh[C.DELIVERED].fillna(0)
    inventory_wh[C.LOG_ID] = np.where(inventory_wh[C.DELIVERED] > 0, log_id, np.nan)
    inventory_wh[C.INVENTORY] = inventory_wh[C.INVENTORY] - inventory_wh[C.DELIVERED]
    updated_inv.append(inventory_wh.drop(columns=[C.DELIVERED]))

def split_ordered_quantity_by_warehouse_codes(po, column):
    # po = po[po["Tienda"] == 27]
    if C.STORE_ID not in po.columns:
        po[C.STORE_ID] = 0
    group_cols = [C.STORE_ID, column]
    po[C.MISSING] = po[C.ORDERED] - po.groupby(group_cols)[C.DELIVERED].transform("sum")

    group_indices = po.groupby(group_cols).cumcount()
    group_sizes = po.groupby(group_cols).transform('size')

    # Identify the last row in each group by comparing group index to size - 1
    po['_is_last'] = group_indices == (group_sizes - 1)

    # Adjust ORDERED based on position
    po[~po['_is_last']][C.ORDERED] = po.loc[~po['_is_last'], C.DELIVERED].values
    po[po['_is_last']][C.ORDERED] = (
        po.loc[po['_is_last'], C.DELIVERED] + po.loc[po['_is_last'], C.MISSING]
    ).values

    # Clean up
    po = po.drop(columns=[C.MISSING, '_is_last'])
    return po

def adjust_quantities_per_row(group):
    group = group.copy()
    if len(group) > 1:
        group.iloc[:-1, group.columns.get_loc(C.ORDERED)] = group[C.DELIVERED].iloc[:-1]
        group.iloc[-1, group.columns.get_loc(C.ORDERED)] = group[C.DELIVERED].iloc[-1] + group[C.MISSING].iloc[-1]
    return group

def update_inventory_in_memory(updated_inv, inventory, log_id, config):
    updated_inv[C.RECEIVED_DATE] = pd.to_datetime(updated_inv[C.RECEIVED_DATE]).dt.date
    # for col in [C.WAREHOUSE_CODE, C.UPC, C.SKU]:
    #     updated_inv[col] = updated_inv[col].astype(int)
    invoc.save_excel(updated_inv, 'INVENTARIO/INVENTARIO.xlsx')
    invoc.save_csv(inventory, f'INVENTARIO/SNAPSHOTS/inventory_{log_id}.csv')
    inv_summ_cols = config.get('inventory_summ_columns')
    inv_summ = updated_inv.groupby(inv_summ_cols)[C.INVENTORY].sum().reset_index()
    invoc.save_excel(inv_summ, 'INVENTARIO/SUMMARY.xlsx')


def save_df_in_excel_and_keep_other_sheets(df, path, sheet_name="Sheet1"):
    # Write new data, replacing the sheet but keeping others
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def sort_rd(rd):
    match = re.match(r'([a-zA-Z])(\d+)([a-zA-Z]?)', rd)
    prefix = match.group(1)
    number = int(match.group(2))
    suffix = match.group(3) or ''
    return number, suffix, prefix

def allocate_stock(po, inventory, column):
    code_lst = po[column].unique()
    delivered = []
    for code in code_lst:
        po_sku = po[po[column] == code]
        inventory_sku = inventory[inventory[column] == code]
        demand = po_sku[C.ORDERED].sum()
        stock = inventory_sku[C.INVENTORY].sum()
        if stock >= demand:
            delivered_sku = po_sku[C.ORDERED]
        else:
            demand_store = po_sku[C.ORDERED].values
            delivered_sku = np.zeros_like(demand_store)
            while stock > 0:
                allocate_i = (demand_store == demand_store.max()).astype(int)
                if allocate_i.sum() >= stock:
                    indices = np.flatnonzero(allocate_i)  # [1, 3, 4]
                    keep_indices = indices[:stock]  # [1, 3]
                    allocate_i = np.zeros_like(allocate_i)
                    allocate_i[keep_indices] = 1
                delivered_sku += allocate_i
                demand_store -= allocate_i
                stock -= allocate_i.sum()
        delivered.append(delivered_sku)
    return np.concatenate(delivered)

def create_and_save_techsmart_txt_file(po, customer, config, po_nums, files_save_path):
    ts_rename = config["ts_rename"]
    ts_columns_txt = config["ts_columns_txt"]
    ts_columns_csv = config["ts_columns_csv"]
    ts = po.copy()
    ts = ts[ts[C.DELIVERED] > 0].reset_index(drop=True)
    ts = ts.rename(columns=ts_rename)
    ts['Tipo'] = np.where(ts['Cantidad'] > 0, 'Salida', 'Entrada')
    ts['Cantidad'] = ts['Cantidad'].abs()
    ts['FECHA'] = date.today().strftime('%d/%m/%Y')
    ts['Cliente final'] = customer.title()
    ts['Unidad'] = 'pzas'
    ts['Caja final'] = ts['Caja inicial']
    add_nan_cols(ts, list(set(ts_columns_txt + ts_columns_csv)))
    # po[(po[STORE_ID] == 688) & (po[SKU] == 1035942641)]
    invoc.save_csv(ts[ts_columns_txt], f"{files_save_path}/techsmart_{str(po_nums)}.txt", sep='\t')
    return ts[ts_columns_csv]

def add_nan_cols(df, cols):
    for col in cols:
        if col not in df.columns:
            df[col] = np.nan

def save_checklist(po_style, po_store, techsmart, config, po_nums, files_save_path):
    cols = config['checklist_columns']
    checklist = po_style[cols]
    dfs = [checklist, po_store, techsmart]
    sheet_names = ["CHECKLIST", "TIENDA", "TECHSMART"]
    checklist_path = f"{files_save_path}/Checklist_{po_nums}.xlsx"
    invoc.save_multiple_dfs_to_excel(dfs, sheet_names, checklist_path, auto_adjust_columns=True)

def autoadjust_column_widths(file_path):
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = max_length + 2  # Padding for readability
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    wb.save(file_path)

def update_billing_record(po_style, customer, delivery_date, config, txn_key, log_id):
    br_columns = config["br_columns"]
    br = invoc.read_excel('FACTURACION/FACTURACION.xlsx')
    po_br = po_style.copy()
    po_br[C.DELIVERY_DATE] = datetime.strptime(delivery_date, "%m/%d/%Y")
    po_br[C.KEY] = txn_key[0] if txn_key else np.nan
    po_br[C.CUSTOMER] = customer.upper()
    po_br[C.INVOICED] = po_br[C.ORDERED]
    po_br[C.SUBTOTAL] = 0 if txn_key else po_br[C.ORDERED] * po_br[C.WHOLESALE_PRICE]
    po_br[C.DISCOUNT] = 0
    if customer.lower() == 'liverpool':
        po_br[C.DISCOUNT] = po_br[C.SUBTOTAL] * .035
    po_br[C.SUBTOTAL_NET] = po_br[C.SUBTOTAL] - po_br[C.DISCOUNT]
    po_br[C.VAT] = po_br[C.SUBTOTAL_NET] * 1.16
    po_br[C.LOG_ID] = log_id
    bru = pd.concat([br, po_br], ignore_index=True)[br_columns]
    bru[C.DELIVERY_DATE] = pd.to_datetime(bru[C.DELIVERY_DATE]).dt.date
    invoc.save_excel(bru, 'FACTURACION/FACTURACION.xlsx')




