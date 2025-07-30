import json
import os
import warnings
from itertools import product

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from src.inventory.common import (create_and_save_techsmart_txt_file, save_checklist, update_billing_record,
                                  add_nan_cols)
from src.inventory.varnames import ColNames as C

def summarize_po_deliveries_by_lot_number(df):
    summary = df.loc[(df[C.ORDERED] > 0),
    [C.RD, C.WAREHOUSE_CODE, C.STYLE, C.UPC, C.SKU, C.ORDERED, C.DELIVERED]]
    summary = summary.reset_index().sort_values(by=['SKU', 'index']).drop('index', axis=1).reset_index(drop=True)
    summary[C.ORDERED] = summary.groupby(C.SKU).apply(
        lambda group: group[C.DELIVERED].where(group.index != group.index[-1], group[C.ORDERED])
    ).reset_index(drop=True)
    summary = summary[summary[C.ORDERED] > 0]
    summary.to_csv('../../files/inventory/po_delivered.csv', index=False)
    return summary

def find_best_carton_combo(total_volume, max_cartons, capacities, costs):
    best_combo = ()
    best_score = float('inf')

    for combo in product(range(max_cartons + 1), repeat=len(capacities)):
        total_cartons = sum(combo)
        if total_cartons > max_cartons:
            continue  # Skip this combo

        combo_volume = sum(capacities[i] * combo[i] for i in range(len(combo)))

        if combo_volume >= total_volume and total_cartons > 0:
            score = sum(costs[i] * combo[i] for i in range(len(combo)))

            if score < best_score:
                best_score = score
                best_combo = tuple(capacities[i] for i in range(len(combo)) for _ in range(combo[i]))
    return best_combo

def assign_box_number(po, rfid_series, cartons):
    names, capacities, costs, dimensions = get_cartons_info(cartons)
    po = assign_box_combos_per_store(po, capacities, costs)

    # Assignments
    stores = po[C.STORE_ID].values
    row_volume = po['ROW_VOLUME'].values
    combo = po['COMBO'].values
    box_assignment = []
    cum_space = []
    store_prev = stores[0]
    rs = 0
    c = 0
    box, end_box = [int(i[1:]) for i in rfid_series[rs]]

    for store_s, space_s, combo_s in zip(stores, row_volume, combo):
        cum_space.append(space_s)
        max_vol = combo_s[c] if c < (len(combo_s) - 1) else capacities[0]
        if ((space_s > 0) & (sum(cum_space) > max_vol)) | (store_s != store_prev):
            box += 1
            c += 1
            cum_space = [space_s]
        if store_s != store_prev:
            c = 0
            store_prev = store_s
        if box > end_box:
            rs += 1
            box, end_box = [int(i[1:]) for i in rfid_series[rs]]
        box_assignment.append(box)

    po = add_box_related_columns(po, box_assignment, names, capacities, dimensions)
    return po


def get_cartons_info(cartons):
    names = [c["name"] for c in cartons]
    capacities = [c["capacity"] for c in cartons]
    costs = [c["cost"] for c in cartons]
    dimensions = [tuple(c["dimensions"]) for c in cartons]
    return names, capacities, costs, dimensions


def add_box_related_columns(po, box_assignment, names, capacities, dimensions):
    po[C.BOX_ID] = ["C" + str(i) for i in box_assignment]
    po['BOX_CHANGE'] = (po[C.BOX_ID] != po[C.BOX_ID].shift()).astype(int)
    po['BOX_STORE_NUM'] = po.groupby([C.STORE_ID])['BOX_CHANGE'].cumsum()
    po['BOX_VOLUME'] = po.groupby(C.BOX_ID)['ROW_VOLUME'].transform('sum')
    po[C.BOX_TYPE] = pd.cut(po['BOX_VOLUME'], bins=[0] + capacities[::-1], labels=names[::-1], right=True).astype(str)
    name_to_dimensions = {n: d for n, d in zip(names, dimensions)}
    po['BOX_DIMENSION'] = po[C.BOX_TYPE].map(name_to_dimensions)
    po[['LENGTH', 'WIDTH', 'HEIGHT']] = pd.DataFrame(po['BOX_DIMENSION'].tolist(), index=po.index)
    return po


def assign_box_combos_per_store(po, capacities, costs):
    po['ROW_VOLUME'] = (capacities[0] / po[C.PCS_BOX]) * po[C.DELIVERED]
    store_volumes = po.groupby([C.STORE_ID])['ROW_VOLUME'].sum().reset_index(name='STORE_VOLUME')
    store_volumes['STORE_VOLUME'] = store_volumes[
                                        "STORE_VOLUME"] * 1.04  # to account that a greedy assignment by row leaves empty space
    store_volumes['MAX_CARTON'] = np.ceil(store_volumes['STORE_VOLUME'] / capacities[2]).astype(int)
    store_volumes['COMBO'] = [find_best_carton_combo(vol, max_crtn, capacities, costs) for
                              vol, max_crtn in zip(store_volumes['STORE_VOLUME'], store_volumes['MAX_CARTON'])]
    po = po.merge(store_volumes[[C.STORE_ID, 'COMBO']], on=[C.STORE_ID], how='left')
    return po

def create_po_summary_by_store(po, config):
    po_gb = po.groupby(config['store_indexes'])[C.DELIVERED].sum().reset_index()
    return po_gb

def update_records_and_save_po_files(po, customer, delivery_date, config, po_path):
    po_num = po.loc[0, C.PO_NUM]
    section = po.loc[0, C.SECTION]
    po_style = create_po_summary_by_style(po, config)
    po_store = create_po_summary_by_store(po, config)
    techsmart = create_and_save_techsmart_txt_file(po, customer, config, po_num, po_path)
    save_checklist(po_style, po_store, techsmart, config, po_num, po_path)
    create_and_save_delivery_note(po_style, delivery_date, config, po_num, section, po_path)
    create_and_save_asn_file(po, config, po_path)
    update_billing_record(po_style, customer, delivery_date, config, "V")



def create_and_save_delivery_note(po_style, delivery_date, config, po_num, section, po_path):
    dn_structure = config["dn_structure"]
    delivery_num = int(dn_structure[0][1]) + 1
    dn_structure[0][1] = delivery_num
    for line, value in zip([0, 8, 9, 10], [delivery_num, po_num, section, delivery_date]):
        dn_structure[line][1] = str(value)
    blank_row = pd.DataFrame([[]])
    dn_columns = config["dn_columns"]
    dn = po_style.groupby(dn_columns[1:5]).agg({
        C.DELIVERED: 'sum', C.CUSTOMER_COST: 'mean'
    }).reset_index()[dn_columns]
    dn['SUBTOTAL'] = dn[C.DELIVERED] * dn[C.CUSTOMER_COST]
    subtotal = dn['SUBTOTAL'].sum()
    discount = subtotal * .045
    subtotal_2 = subtotal - discount
    vat = subtotal_2 * .016
    total = subtotal_2 + vat
    dn_totals = pd.DataFrame({5: ["Subtotal", "Descuento 4.5%", "SubTotal Menos", "IVA", "Total"],
                              6: [subtotal, discount, subtotal_2, vat, total], })
    delivery_note = pd.concat([pd.DataFrame(dn_structure), blank_row, dn.T.reset_index().T, dn_totals],
                              ignore_index=True)
    dn_file_path = f"../../files/inventory/{po_path}/Nota_Remision_{delivery_num}.xlsx"
    delivery_note.to_excel(dn_file_path, index=False, header=False, sheet_name='Sheet1')
    format_delivery_note(dn_file_path)


def format_delivery_note(dn_file_path):
    wb = load_workbook(dn_file_path)
    ws = wb['Sheet1']
    col_widths = [22, 20, 50, 15, 15, 15, 20]
    for col_letter, width in zip(list('ABCDEFG'), col_widths):  # start=1 for column A
        ws.column_dimensions[col_letter].width = width
    # === Fit to A4 page when printing ===
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # or PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1  # Fit to 1 page wide
    ws.page_setup.fitToHeight = 0  # Unlimited height (or set 1 to fit in 1 page tall)
    # Set margins
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    for row in range(14, ws.max_row + 1):  # Start at row 2 if row 1 is header
        cell = ws[f"{'G'}{row}"]
        cell.number_format = '#,##0.00'  # Format without decimals
    # Save final Excel file
    wb.save(dn_file_path)

def create_and_save_asn_file(po, config, po_path):
    asn_rename = config["asn_rename"]
    asn_columns = config["asn_columns"]
    asn = po.copy().rename(columns=asn_rename)
    asn['CENTRO/ALMACEN DESTINO'] = [f"{c:04}" for c in asn['CENTRO']]
    asn['TRANSPORTE'] = 1
    asn[C.CUSTOMER_UPC] = np.nan
    add_nan_cols(asn, asn_columns)
    asn[asn_columns].to_excel(f"../../files/inventory/{po_path}/asn.xlsx", index=False)

def create_po_summary_by_style(po, config):
    po_gb = po.groupby(config['checklist_indexes']).agg(config['checklist_values']).reset_index()
    return po_gb

def run_process_purchase_orders(po, config, customer, delivery_date, po_path, rfid_series=None):
    po = po[(~po[C.STYLE].isna())].reset_index(drop=True)
    styles_pc = po.loc[(po[C.CUSTOMER_COST] != po[C.WHOLESALE_PRICE]), C.CUSTOMER_STYLE].unique() #pc = price_conflict
    if len(styles_pc) > 0:
        warnings.warn(f"""The following styles have price conflicts: {", ".join(styles_pc)}""", UserWarning)
    po = assign_box_number(po, rfid_series, config["cartons"])
    update_records_and_save_po_files(po, customer, delivery_date, config, po_path)
    with open("../../files/inventory/config_vars.json", "w") as f:
        json.dump(config, f, indent=2)


# new_inventory = update_inventory_in_memory(new_inventory_dfs)
# po_summary = summarize_po_deliveries_by_lot_number(new_inventory)
    # inventory = inventory[inventory[SKU] == 1145889908]
    # po = po[po[SKU] == 1035942579]
    # po = po[po[C.SKU] == 1035942641] # split

    x=1