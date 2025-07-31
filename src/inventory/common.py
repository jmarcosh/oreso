import json
import os
import re

import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from openpyxl import load_workbook

from src.inventory.varnames import ColNames as C
from src.api_integrations.sharepoint_client import SharePointClient
invoc = SharePointClient()

def get_all_xlsx_files_in_directory(directory_path):
    if not os.path.isdir(directory_path):
        raise FileNotFoundError(f"{directory_path} is not a valid directory.")

    return [
        os.path.join(directory_path, f)
        for f in os.listdir(directory_path)
        if os.path.isfile(os.path.join(directory_path, f)) and f.lower().endswith('.xlsx')
    ]


def read_files_and_backup_inventory(log_id):
    config = invoc.read_json("config/config.json")
    inventory_df = invoc.read_excel('INVENTARIO/INVENTARIO.xlsx', sheet_name='Data')
    invoc.save_csv(inventory_df, f'INVENTARIO/prev_snapshots/inventory_{log_id}.csv',)
    po_read_path = '../../files/inventory/drag_and_drop'
    po_files = get_all_xlsx_files_in_directory(po_read_path)
    po_dfs = [pd.read_excel(po_path, sheet_name=0) for po_path in po_files]
    po_df = pd.concat(po_dfs)
    customer, matching_column = auto_assign_customer(po_df)
    cols_rename = config[f'{customer.lower()}_rename']
    cols = cols_rename.values()
    return (po_df.rename(cols_rename, axis=1).sort_values(by=C.SKU)[cols].reset_index(drop=True), inventory_df, config,
            customer, matching_column)


def auto_assign_customer(df):
    if '# Prov' in df.columns:
        return 'liverpool', C.SKU
    elif 'Num. Prov' in df.columns:
        return 'suburbia', C.SKU
    return 'interno', C.WAREHOUSE_CODE


def assign_warehouse_codes_from_column_and_update_inventory(po, inventory, column):
    print(inventory[C.INVENTORY].sum())
    split_inventory = split_df_by_column(inventory, column)
    po_missing = po.loc[(po[C.DELIVERED] == 0)].merge(
            split_inventory[1],
            on=[column], how='left')
    po_original_cols = po.columns
    it = 0
    po_wh = [po_missing]
    updated_inv = [split_inventory[0]]
    for inventory_wh in split_inventory[1:]:
        it += 1
        po = split_deliveries_by_column(po, inventory_wh, column, po_original_cols, po_wh, updated_inv)
        if len(po) == 0:
            break
    po = pd.concat(po_wh)
    po = split_ordered_quantity_by_stores(po, column)
    updated_inv = updated_inv + split_inventory[it+1:]
    update_inventory_in_memory(updated_inv)
    return po.sort_values([C.STORE_ID, C.SKU]).reset_index(drop=True)


def split_deliveries_by_column(po, inventory_wh, column, po_original_cols, po_wh, updated_inv):
    po = po.merge(
        inventory_wh,
        on=[column], how='left')
    delivered_cs = po.groupby(column)[C.DELIVERED].cumsum()
    missing = (po[C.INVENTORY] - delivered_cs).clip(upper=0).fillna(0)
    delivered_i = (po[C.DELIVERED] + missing).clip(lower=0)
    to_deliver = po[C.DELIVERED] - delivered_i
    po[C.DELIVERED] = delivered_i
    po_wh.append(po.loc[po[C.DELIVERED] > 0].copy())
    update_inventory(inventory_wh, po, updated_inv)
    po[C.DELIVERED] = to_deliver
    po = po.loc[po[C.DELIVERED] > 0, po_original_cols]
    return po


def split_df_by_column(df, column):
    column0 = df[df[column] == 0]
    df = df.drop(index=column0.index)
    split_dfs = [column0]
    while not df.empty:
        unique_df = df.drop_duplicates(subset=[column], keep='first')
        split_dfs.append(unique_df)
        df = df.drop(index=unique_df.index)
    return split_dfs

def update_inventory(inventory_wh, po, updated_inv):
    inventory_wh = inventory_wh.join(po.groupby([C.WAREHOUSE_CODE])[C.DELIVERED].sum(), on=C.WAREHOUSE_CODE)
    inventory_wh[C.DELIVERED] = inventory_wh[C.DELIVERED].fillna(0)
    inventory_wh[C.INVENTORY] = inventory_wh[C.INVENTORY] - inventory_wh[C.DELIVERED]
    updated_inv.append(inventory_wh.drop(columns=[C.DELIVERED]))

def split_ordered_quantity_by_stores(po, column):
    # po = po[po["Tienda"] == 27]
    po[C.MISSING] = po[C.ORDERED] - po.groupby([C.STORE_ID, column])[C.DELIVERED].transform("sum")
    po = po.groupby([C.STORE_ID, column], group_keys=False).apply(adjust_quantities_per_row)
    return po.drop(columns=[C.MISSING])

def adjust_quantities_per_row(group):
    group = group.copy()
    if len(group) > 1:
        group.iloc[:-1, group.columns.get_loc(C.ORDERED)] = group[C.DELIVERED].iloc[:-1]
        group.iloc[-1, group.columns.get_loc(C.ORDERED)] = group[C.DELIVERED].iloc[-1] + group[C.MISSING].iloc[-1]
    return group

def update_inventory_in_memory(dfs):
    df = pd.concat(dfs)
    # df = df.sort_values(
    #     by=['RD', 'WAREHOUSE_CODE'],
    #     key=lambda col: col.map(sort_rd) if col.name == 'code' else col,
    #     ignore_index=True
    # )
    df.sort_index(inplace=True)
    # df = df[df[C.INVENTORY] > 0]
    df[C.RECEIVED_DATE] = pd.to_datetime(df[C.RECEIVED_DATE]).dt.date
    invoc.save_excel(df, 'INVENTARIO/INVENTARIO_U.xlsx')

def save_df_in_excel_and_keep_other_sheets(df, path, sheet_name="Data"):
    # Write new data, replacing the sheet but keeping others
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def sort_rd(rd):
    match = re.match(r'([a-zA-Z])(\d+)([a-zA-Z]?)', rd)
    prefix = match.group(1)
    number = int(match.group(2))
    suffix = match.group(3) or ''
    return number, suffix, prefix

def allocate_stock(po, inventory, column):
    code_lst = po[column].unique()
    delivered = []
    for code in code_lst:
        po_sku = po[po[column] == code]
        inventory_sku = inventory[inventory[column] == code]
        demand = po_sku[C.ORDERED].sum()
        stock = inventory_sku[C.INVENTORY].sum()
        if stock >= demand:
            delivered_sku = po_sku[C.ORDERED]
        else:
            demand_store = po_sku[C.ORDERED].values
            delivered_sku = np.zeros_like(demand_store)
            while stock > 0:
                allocate_i = (demand_store == demand_store.max()).astype(int)
                if allocate_i.sum() >= stock:
                    indices = np.flatnonzero(allocate_i)  # [1, 3, 4]
                    keep_indices = indices[:stock]  # [1, 3]
                    allocate_i = np.zeros_like(allocate_i)
                    allocate_i[keep_indices] = 1
                delivered_sku += allocate_i
                demand_store -= allocate_i
                stock -= allocate_i.sum()
        delivered.append(delivered_sku)
    return np.concatenate(delivered)

def create_and_save_techsmart_txt_file(po, customer, config, po_num, files_save_path):
    ts_rename = config["ts_rename"]
    ts_columns_txt = config["ts_columns_txt"]
    ts_columns_csv = config["ts_columns_csv"]
    ts = po.copy()
    ts = ts[ts[C.DELIVERED] > 0].reset_index(drop=True)
    ts = ts.rename(columns=ts_rename)
    ts['Tipo'] = np.where(ts['Cantidad'] > 0, 'Salida', 'Entrada')
    ts['Cantidad'] = ts['Cantidad'].abs()
    ts['FECHA'] = date.today().strftime('%d/%m/%Y')
    ts['Cliente final'] = customer.title()
    ts['Unidad'] = 'pzas'
    ts['Caja final'] = ts['Caja inicial']
    add_nan_cols(ts, list(set(ts_columns_txt + ts_columns_csv)))
    # po[(po[STORE_ID] == 688) & (po[SKU] == 1035942641)]
    invoc.save_csv(ts[ts_columns_txt], f"{files_save_path}/techsmart_{str(po_num)}.txt", sep='\t')
    return ts[ts_columns_csv]

def add_nan_cols(df, cols):
    for col in cols:
        if col not in df.columns:
            df[col] = np.nan

def save_checklist(po_style, po_store, techsmart, config, po_num, files_save_path):
    checklist_cols = [x for x in config['checklist_indexes'] if x not in [C.PO_NUM, C.GROUP]]
    checklist = po_style.groupby(checklist_cols)[[C.INVENTORY, C.ORDERED, C.DELIVERED]].sum().reset_index().sort_values(by=[C.SKU])
    dfs = [checklist, po_store, techsmart]
    sheet_names = ["CHECKLIST", "TIENDA", "TECHSMART"]
    checklist_path = f"{files_save_path}/Checklist_{po_num}.xlsx"
    invoc.save_multiple_dfs_to_excel(dfs, sheet_names, checklist_path, auto_adjust_columns=True)

def autoadjust_column_widths(file_path):
    wb = load_workbook(file_path)
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = max_length + 2  # Padding for readability
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    wb.save(file_path)

def update_billing_record(po_style, customer, delivery_date, config, txn_key=None):
    br_columns = config["br_columns"]
    br = invoc.read_excel('FACTURACION/FACTURACION.xlsx', sheet_name="Data")
    po_br = po_style.copy()
    po_br[C.DELIVERY_DATE] = datetime.strptime(delivery_date, "%m/%d/%Y")
    po_br[C.SHIPPED] = 'TULTITLAN'
    po_br[C.KEY] = txn_key[0] if txn_key else np.nan
    po_br[C.CUSTOMER] = customer.upper()
    po_br[C.INVOICE_NUM] = txn_key if txn_key else np.nan
    po_br[C.SUBTOTAL] = 0 if txn_key else po_br[C.ORDERED] * po_br[C.WHOLESALE_PRICE]
    po_br[C.DISCOUNT] = 0
    if customer.lower() == 'liverpool':
        po_br[C.DISCOUNT] = po_br[C.SUBTOTAL] * .035
    po_br[C.SUBTOTAL_NET] = po_br[C.SUBTOTAL] - po_br[C.DISCOUNT]
    po_br[C.VAT] = po_br[C.SUBTOTAL_NET] * 1.16
    bru = pd.concat([br, po_br], ignore_index=True)[br_columns]
    for col in [C.COLLECTED, C.DELIVERY_DATE, C.INVOICE_DATE]:
        bru[col] = pd.to_datetime(bru[col]).dt.date
    invoc.save_excel(bru, 'FACTURACION/FACTURACION_u.xlsx')


def create_po_path_for_savings(customer, delivery_date, po_num):
    folder_path = f"{customer}/{delivery_date.split('/')[2]}/{delivery_date.split('/')[0]}/{str(po_num)}"
    invoc.create_folder_path("OC", folder_path)
    return "OC/" + folder_path