import json
import warnings
from itertools import product
import re

import numpy as np
import pandas as pd
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from src.inventory.varnames import ColNames as C



def read_files_and_backup_inventory(config, customer):
    inventory_df = pd.read_excel('../../files/inventory/inventory.xlsx', sheet_name='Data')
    today_date = date.today().strftime('%Y%m%d-%H%M')
    # inventory_df.to_csv('../../files/inventory/inventory_' + today_date + '.csv', index=False)
    # po_paths = ['../../files/inventory/purchase_order.xlsx']
    po_paths = ['../../files/inventory/80-013401-20250716_048343-001-ORD_COMPRA.xlsx']
    po_dfs = [pd.read_excel(po_path, sheet_name=0) for po_path in po_paths]
    po_df = pd.concat(po_dfs)
    cols_rename = config[f'{customer.lower()}_rename']
    cols = cols_rename.values()
    return po_df.rename(cols_rename, axis=1).sort_values(by=C.SKU)[cols], inventory_df


def split_df_by_sku(df):
    sku0 = df[df[C.SKU] == 0]
    df = df.drop(index=sku0.index)
    split_dfs = [sku0]
    while not df.empty:
        unique_df = df.drop_duplicates(subset=[C.SKU], keep='first')
        split_dfs.append(unique_df)
        df = df.drop(index=unique_df.index)
    return split_dfs

def save_df_in_excel_and_keep_other_sheets(df, path, sheet_name="Data"):
    # Write new data, replacing the sheet but keeping others
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

def sort_rd(rd):
    match = re.match(r'([a-zA-Z])(\d+)([a-zA-Z]?)', rd)
    prefix = match.group(1)
    number = int(match.group(2))
    suffix = match.group(3) or ''
    return number, suffix, prefix

def update_inventory_in_memory(dfs):
    df = pd.concat(dfs)
    print(df[C.INVENTORY].sum())
    # df = df.sort_values(
    #     by=['RD', 'WAREHOUSE_CODE'],
    #     key=lambda col: col.map(sort_rd) if col.name == 'code' else col,
    #     ignore_index=True
    # )
    df.sort_index(inplace=True)
    df = df[df[C.INVENTORY] > 0]
    df[C.RECEIVED_DATE] = pd.to_datetime(df[C.RECEIVED_DATE]).dt.date
    save_df_in_excel_and_keep_other_sheets(df, '../../files/inventory/inventory_u.xlsx')
    # df.to_excel('../../files/inventory/inventory_u.xlsx', index=False)
    # df_grouped = df.groupby(['SKU', 'UPC']).agg({INVENTORY: 'sum'}).reset_index()
    # df_grouped.to_excel('../../files/inventory/inventory_grouped.xlsx', index=False)
    # return df


def summarize_po_deliveries_by_lot_number(df):
    summary = df.loc[(df[C.ORDERED] > 0),
    [C.RD, C.WAREHOUSE_CODE, C.STYLE, C.UPC, C.SKU, C.ORDERED, C.DELIVERED]]
    summary = summary.reset_index().sort_values(by=['SKU', 'index']).drop('index', axis=1).reset_index(drop=True)
    summary[C.ORDERED] = summary.groupby(C.SKU).apply(
        lambda group: group[C.DELIVERED].where(group.index != group.index[-1], group[C.ORDERED])
    ).reset_index(drop=True)
    summary = summary[summary[C.ORDERED] > 0]
    summary.to_csv('../../files/inventory/po_delivered.csv', index=False)
    return summary

def allocate_stock(po, inventory):
    sku_lst = po[C.SKU].unique()
    delivered = []
    for sku in sku_lst:
        po_sku = po[po[C.SKU] == sku]
        inventory_sku = inventory[inventory[C.SKU] == sku]
        demand = po_sku[C.ORDERED].sum()
        stock = inventory_sku[C.INVENTORY].sum()
        if stock >= demand:
            delivered_sku = po_sku[C.ORDERED]
        else:
            demand_store = po_sku[C.ORDERED].values
            delivered_sku = np.zeros_like(demand_store)
            while stock > 0:
                allocate_i = (demand_store == demand_store.max()).astype(int)
                if allocate_i.sum() >= stock:
                    indices = np.flatnonzero(allocate_i)  # [1, 3, 4]
                    keep_indices = indices[:stock]  # [1, 3]
                    allocate_i = np.zeros_like(allocate_i)
                    allocate_i[keep_indices] = 1
                delivered_sku += allocate_i
                demand_store -= allocate_i
                stock -= allocate_i.sum()
        delivered.append(delivered_sku)
    return np.concatenate(delivered)

def assign_warehouse_codes_from_sku_and_update_inventory(po, inventory):
    print(inventory[C.INVENTORY].sum())
    split_inventory = split_df_by_sku(inventory[inventory[C.INVENTORY] > 0])
    po_missing = po.loc[(po[C.DELIVERED] == 0)].merge(
            split_inventory[1],
            on=[C.SKU], how='left')
    po_wh = [po_missing]
    po_original_cols = po.columns
    it = 0
    # split_inventory[0]['ROUND'] = it
    # split_inventory[0]['INVENTORY_UP'] = split_inventory[0][INVENTORY]
    updated_inv = [split_inventory[0]]
    for inventory_wh in split_inventory[1:]:
        it += 1
        po = po.merge(
            inventory_wh,
            on=[C.SKU], how='left')
        delivered_cs = po.groupby(C.SKU)[C.DELIVERED].cumsum()
        missing = (po[C.INVENTORY] - delivered_cs).clip(upper=0).fillna(0)
        delivered_i = (po[C.DELIVERED] + missing).clip(lower=0)
        to_deliver = po[C.DELIVERED] - delivered_i
        po[C.DELIVERED] = delivered_i
        po_wh.append(po.loc[po[C.DELIVERED] > 0].copy())
        update_inventory(inventory_wh, po, updated_inv)
        po[C.DELIVERED] = to_deliver
        po = po.loc[po[C.DELIVERED] > 0, po_original_cols]
        if len(po) == 0:
            break
    po = pd.concat(po_wh)
    po = split_ordered_quantity_into_warehouse_codes(po)
    print(po[C.DELIVERED].sum())
    updated_inv = updated_inv + split_inventory[it+1:]
    update_inventory_in_memory(updated_inv)
    return po.sort_values([C.STORE_ID, C.SKU]).reset_index(drop=True)


def update_inventory(inventory_wh, po, updated_inv):
    inventory_wh = inventory_wh.join(po.groupby([C.WAREHOUSE_CODE])[C.DELIVERED].sum(), on=C.WAREHOUSE_CODE)
    inventory_wh[C.DELIVERED] = inventory_wh[C.DELIVERED].fillna(0)
    inventory_wh[C.INVENTORY] = inventory_wh[C.INVENTORY] - inventory_wh[C.DELIVERED]
    updated_inv.append(inventory_wh.drop(columns=[C.DELIVERED]))


def adjust_quantities_per_row(group):
    group = group.copy()
    if len(group) > 1:
        group.iloc[:-1, group.columns.get_loc(C.ORDERED)] = group[C.DELIVERED].iloc[:-1]
        group.iloc[-1, group.columns.get_loc(C.ORDERED)] = group[C.DELIVERED].iloc[-1] + group[C.MISSING].iloc[-1]
    return group

def split_ordered_quantity_into_warehouse_codes(po):
    # po = po[po["Tienda"] == 27]
    po[C.MISSING] = po[C.ORDERED] - po.groupby([C.STORE_ID, C.SKU])[C.DELIVERED].transform("sum")
    po = po.groupby([C.STORE_ID, C.SKU], group_keys=False).apply(adjust_quantities_per_row)
    return po.drop(columns=[C.MISSING])

def find_best_carton_combo(total_volume, max_cartons, capacities, costs):
    best_combo = ()
    best_score = float('inf')

    for combo in product(range(max_cartons + 1), repeat=len(capacities)):
        total_cartons = sum(combo)
        if total_cartons > max_cartons:
            continue  # Skip this combo

        combo_volume = sum(capacities[i] * combo[i] for i in range(len(combo)))

        if combo_volume >= total_volume and total_cartons > 0:
            score = sum(costs[i] * combo[i] for i in range(len(combo)))

            if score < best_score:
                best_score = score
                best_combo = tuple(capacities[i] for i in range(len(combo)) for _ in range(combo[i]))
    return best_combo


def assign_box_number(po, rfid_series, cartons):
    names, capacities, costs, dimensions = get_cartons_info(cartons)
    po = assign_box_combos_per_store(po, capacities, costs)

    # Assignments
    stores = po[C.STORE_ID].values
    row_volume = po['ROW_VOLUME'].values
    combo = po['COMBO'].values
    box_assignment = []
    cum_space = []
    store_prev = stores[0]
    rs = 0
    c = 0
    box, end_box = [int(i[1:]) for i in rfid_series[rs]]

    for store_s, space_s, combo_s in zip(stores, row_volume, combo):
        cum_space.append(space_s)
        max_vol = combo_s[c] if c < (len(combo_s) - 1) else capacities[0]
        if (space_s > 0) & ((sum(cum_space) > max_vol) | (store_s != store_prev)) :
            box += 1
            c += 1
            cum_space = [space_s]
        if store_s != store_prev:
            c = 0
            store_prev = store_s
        if box > end_box:
            rs += 1
            box, end_box = [int(i[1:]) for i in rfid_series[rs]]
        box_assignment.append(box)

    po = add_box_related_columns(po, box_assignment, names, capacities, dimensions)
    return po


def get_cartons_info(cartons):
    names = [c["name"] for c in cartons]
    capacities = [c["capacity"] for c in cartons]
    costs = [c["cost"] for c in cartons]
    dimensions = [tuple(c["dimensions"]) for c in cartons]
    return names, capacities, costs, dimensions


def add_box_related_columns(po, box_assignment, names, capacities, dimensions):
    po[C.BOX_ID] = ["C" + str(i) for i in box_assignment]
    po['BOX_CHANGE'] = (po[C.BOX_ID] != po[C.BOX_ID].shift()).astype(int)
    po['BOX_STORE_NUM'] = po.groupby([C.STORE_ID])['BOX_CHANGE'].cumsum()
    po['BOX_VOLUME'] = po.groupby(C.BOX_ID)['ROW_VOLUME'].transform('sum')
    po[C.BOX_TYPE] = pd.cut(po['BOX_VOLUME'], bins=[0] + capacities[::-1], labels=names[::-1], right=True).astype(str)
    name_to_dimensions = {n: d for n, d in zip(names, dimensions)}
    po['BOX_DIMENSION'] = po[C.BOX_TYPE].map(name_to_dimensions)
    po[['LENGTH', 'WIDTH', 'HEIGHT']] = pd.DataFrame(po['BOX_DIMENSION'].tolist(), index=po.index)
    return po


def assign_box_combos_per_store(po, capacities, costs):
    po['ROW_VOLUME'] = (capacities[0] / po[C.PCS_BOX]) * po[C.DELIVERED]
    store_volumes = po.groupby([C.STORE_ID])['ROW_VOLUME'].sum().reset_index(name='STORE_VOLUME')
    store_volumes['STORE_VOLUME'] = store_volumes[
                                        "STORE_VOLUME"] * 1.04  # to account that a greedy assignment by row leaves empty space
    store_volumes['MAX_CARTON'] = np.ceil(store_volumes['STORE_VOLUME'] / capacities[2]).astype(int)
    store_volumes['COMBO'] = [find_best_carton_combo(vol, max_crtn, capacities, costs) for
                              vol, max_crtn in zip(store_volumes['STORE_VOLUME'], store_volumes['MAX_CARTON'])]
    po = po.merge(store_volumes[[C.STORE_ID, 'COMBO']], on=[C.STORE_ID], how='left')
    return po


def create_and_save_techsmart_txt_file(po, customer, config):
    ts_rename = config["ts_rename"]
    ts_columns_txt = config["ts_columns_txt"] 
    ts_columns_csv = config["ts_columns_csv"]
    ts = po.copy()
    ts = ts[ts[C.DELIVERED] > 0].reset_index(drop=True)
    ts = ts.rename(columns=ts_rename)
    ts['Tipo'] = 'Salida'
    ts['FECHA'] = date.today().strftime('%Y%m%d')
    ts['Cliente final'] = customer
    ts['Unidad'] = 'pzas'
    ts['Caja final'] = ts['Caja inicial']
    add_nan_cols(ts, ts_columns_txt)
    # po[(po[STORE_ID] == 688) & (po[SKU] == 1035942641)]
    ts[ts_columns_txt].to_csv("../../files/inventory/techsmart_" + str(ts.loc[0, '# OC']) + ".txt", sep='\t',
                              index=False)
    return ts[ts_columns_csv]


def create_po_summary_by_store(po):
    po_gb = po.groupby([C.STORE_ID, C.BOX_ID, C.BOX_TYPE])[C.DELIVERED].sum().reset_index()
    return po_gb


def save_checklist(po_style, po_store, techsmart, config):
    checklist_cols = [x for x in po_style.columns[:10] if x not in [C.PO_NUM, C.GROUP]]
    checklist = po_style.groupby(checklist_cols)[[C.INVENTORY, C.ORDERED, C.DELIVERED]].sum().reset_index().sort_values(by=[C.SKU])
    checklist_path = f"../../files/inventory/Checklist_{config[C.PO_NUM]}.xlsx"
    with pd.ExcelWriter(checklist_path, engine="openpyxl") as writer:
        checklist.to_excel(writer, sheet_name="CHECKLIST", index=False)
        po_store.to_excel(writer, sheet_name="TIENDA", index=False)
        techsmart.to_excel(writer, sheet_name="OUTPUT", index=False)
    autoadjust_column_widths(checklist_path, ["CHECKLIST", "TIENDA", "OUTPUT"])


def autoadjust_column_widths(file_path, sheet_names):
    wb = load_workbook(file_path)
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = max_length + 2  # Padding for readability
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    wb.save(file_path)


def update_records_and_save_po_files(po, customer, delivery_date, config):
    for key in [C.PO_NUM, C.SECTION]:
        config[key] = po.loc[0, key]
    po_style = create_po_summary_by_style(po)
    po_store = create_po_summary_by_store(po)
    techsmart = create_and_save_techsmart_txt_file(po, customer, config)
    save_checklist(po_style, po_store, techsmart, config)
    create_and_save_delivery_note(po_style, delivery_date, config)
    create_and_save_asn_file(po, config)
    update_billing_record(po_style, customer, delivery_date, config)



def create_and_save_delivery_note(po_style, delivery_date, config):
    dn_structure = config["dn_structure"]
    delivery_num = int(dn_structure[0][1]) + 1
    dn_structure[0][1] = delivery_num
    for line, value in zip([0, 8, 9, 10], [delivery_num, config[C.PO_NUM], config[C.SECTION], delivery_date]):
        dn_structure[line][1] = str(value)
    blank_row = pd.DataFrame([[]])
    dn_columns = config["dn_columns"]
    dn = po_style.groupby(dn_columns[1:5]).agg({
        C.DELIVERED: 'sum', C.CUSTOMER_COST: 'mean'
    }).reset_index()[dn_columns]
    dn['SUBTOTAL'] = dn[C.DELIVERED] * dn[C.CUSTOMER_COST]
    subtotal = dn['SUBTOTAL'].sum()
    discount = subtotal * .045
    subtotal_2 = subtotal - discount
    vat = subtotal_2 * .016
    total = subtotal_2 + vat
    dn_totals = pd.DataFrame({5: ["Subtotal", "Descuento 4.5%", "SubTotal Menos", "IVA", "Total"],
                              6: [subtotal, discount, subtotal_2, vat, total], })
    delivery_note = pd.concat([pd.DataFrame(dn_structure), blank_row, dn.T.reset_index().T, dn_totals],
                              ignore_index=True)
    dn_file_path = f'../../files/inventory/Nota_Remision_{delivery_num}.xlsx'
    delivery_note.to_excel(dn_file_path, index=False, header=False, sheet_name='Sheet1')
    format_delivery_note(dn_file_path)


def format_delivery_note(dn_file_path):
    wb = load_workbook(dn_file_path)
    ws = wb['Sheet1']
    col_widths = [22, 20, 50, 15, 15, 15, 20]
    for col_letter, width in zip(list('ABCDEFG'), col_widths):  # start=1 for column A
        ws.column_dimensions[col_letter].width = width
    # === Fit to A4 page when printing ===
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE  # or PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1  # Fit to 1 page wide
    ws.page_setup.fitToHeight = 0  # Unlimited height (or set 1 to fit in 1 page tall)
    # Set margins
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    for row in range(14, ws.max_row + 1):  # Start at row 2 if row 1 is header
        cell = ws[f"{'G'}{row}"]
        cell.number_format = '#,##0.00'  # Format without decimals
    # Save final Excel file
    wb.save(dn_file_path)

def create_and_save_asn_file(po, config):
    asn_rename = config["asn_rename"]
    asn_columns = config["asn_columns"]
    asn = po.copy().rename(columns=asn_rename)
    asn['CENTRO/ALMACEN DESTINO'] = [f"{c:04}" for c in asn['CENTRO']]
    asn['TRANSPORTE'] = 1
    add_nan_cols(asn, asn_columns)
    asn[asn_columns].to_excel('../../files/inventory/asn.xlsx', index=False)


def add_nan_cols(df, cols):
    for col in cols:
        if col not in df.columns:
            df[col] = np.nan


def update_billing_record(po_style, customer, delivery_date, config):
    br_columns = config["br_columns"]
    br = pd.read_excel('../../files/inventory/Facturación 2025.xlsx', sheet_name="Data")
    po_br = po_style.copy()
    po_br[C.DELIVERY_DATE] = datetime.strptime(delivery_date, "%m/%d/%Y")
    po_br[C.SHIPPED] = 'TULTITLAN'
    po_br[C.KEY] = "V"
    po_br[C.CUSTOMER] = customer.upper()
    po_br[C.SUBTOTAL] = po_br[C.ORDERED] * po_br[C.WHOLESALE_PRICE]
    po_br[C.DISCOUNT] = 0
    if customer.lower() == 'liverpool':
        po_br[C.DISCOUNT] = po_br[C.SUBTOTAL] * .035
    po_br[C.SUBTOTAL_NET] = po_br[C.SUBTOTAL] - po_br[C.DISCOUNT]
    po_br[C.VAT] = po_br[C.SUBTOTAL_NET] * 1.16
    bru = pd.concat([br, po_br], ignore_index=True)[br_columns]
    for col in [C.COLLECTED, C.DELIVERY_DATE, C.INVOICE_DATE]:
        bru[col] = pd.to_datetime(bru[col]).dt.date
    # save_df_in_excel_and_keep_other_sheets(bru, '../../files/inventory/Facturación 2025_u.xlsx')

def create_po_summary_by_style(po):
    po_gb = po.groupby([C.RD, C.MOVEX_PO, C.PO_NUM, C.WAREHOUSE_CODE, C.STYLE, C.DESCRIPTION,
                          C.UPC, C.SKU, C.BRAND, C.GROUP]).agg({
        C.INVENTORY: 'mean',
        C.ORDERED: 'sum',
        C.DELIVERED: 'sum',
        C.WHOLESALE_PRICE: 'mean',
        C.CUSTOMER_COST: 'mean',
    }).reset_index()
    return po_gb

def run_process_purchase_orders(customer, rfid_series, delivery_date):
    with open("../../files/inventory/config_vars.json", "r") as f:
        config = json.load(f)
    po, inventory = read_files_and_backup_inventory(config, customer)
    # inventory = inventory[inventory[SKU] == 1145889908]
    # po = po[po[SKU] == 1035942579]
    # po = po[po[C.SKU] == 1035942641] # split
    po[C.DELIVERED] = allocate_stock(po, inventory)
    po = assign_warehouse_codes_from_sku_and_update_inventory(po, inventory)
    styles_pc = po.loc[(po[C.CUSTOMER_COST] != po[C.WHOLESALE_PRICE]), C.CUSTOMER_STYLE].unique() #pc = price_conflict
    if len(styles_pc) > 0:
        warnings.warn(f"""The following styles have price conflicts: {", ".join(styles_pc)}""", UserWarning)

    po = assign_box_number(po, rfid_series, config["cartons"])
    # po.to_csv('/home/jmarcosh/Downloads/inv_deb.csv')

    update_records_and_save_po_files(po, customer, delivery_date, config)
    for key in [C.PO_NUM, C.SECTION]:
        config.pop(key, None)
    with open("../../files/inventory/config_vars.json", "w") as f:
        json.dump(config, f, indent=2)


    x=1

if __name__ == '__main__':
    CUSTOMER = 'Suburbia'
    RFID_SERIES = [['C52762712', 'C52762891'],
                   ['C52763895', 'C52764000']]
    DELIVERY_DATE = "1/16/2025"
    run_process_purchase_orders(CUSTOMER, RFID_SERIES, DELIVERY_DATE)
# new_inventory = update_inventory_in_memory(new_inventory_dfs)
# po_summary = summarize_po_deliveries_by_lot_number(new_inventory)

    x=1